#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Veo3Free - AI生成工具 PyWebview + React 版本
"""

import asyncio
import json
import os
import sys
import base64
import io
import subprocess
import platform
import threading
import webbrowser
from datetime import datetime
from pathlib import Path

# Windows 下设置输出编码为 UTF-8
if sys.platform == 'win32':
    if sys.stdout is not None and hasattr(sys.stdout, 'buffer'):
        sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
    if sys.stderr is not None and hasattr(sys.stderr, 'buffer'):
        sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8')

    # 强制 pywebview 使用 WinRT 后端（避免 pythonnet/WinForms 依赖）
    os.environ['PYWEBVIEW_BACKEND'] = 'winrt'

from PIL import Image

try:
    from loguru import logger
except ImportError:
    print("请安装 loguru: pip install loguru")
    sys.exit(1)

try:
    from openpyxl import Workbook, load_workbook
except ImportError:
    print("请安装 openpyxl: pip install openpyxl")
    Workbook = None
    load_workbook = None

try:
    import webview
except ImportError:
    print("请安装 pywebview: pip install pywebview")
    sys.exit(1)

from version import get_version
from updater import check_for_updates, open_download_page
from api_server_fastapi import OpenAIAPICompatServer, API_SERVER_PORT
import api_server_fastapi as api_server_module

try:
    from PJYSDK import PJYSDK
    PJYSDK_AVAILABLE = True
except ImportError:
    PJYSDK_AVAILABLE = False
    logger.warning("PJYSDK 未安装，API 验证功能不可用")

# 确定输出目录位置
if getattr(sys, 'frozen', False):
    # 打包后，使用用户文档目录
    OUTPUT_DIR = Path.home() / "Documents" / "veo3free" / "output"
else:
    # 开发模式，使用项目目录
    OUTPUT_DIR = Path("output")

OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

# 配置loguru日志
LOGS_DIR = OUTPUT_DIR.parent / "logs" if getattr(sys, 'frozen', False) else Path("logs")
LOGS_DIR.mkdir(parents=True, exist_ok=True)

# 获取版本号用于日志
from version import __version__ as APP_VERSION

logger.remove()  # 移除默认handler

# 日志格式：时间 | 版本 | 文件:行号 | 消息
LOG_FORMAT_CONSOLE = "<green>{time:HH:mm:ss}</green> | <cyan>v{extra[ver]}</cyan> | <level>{message}</level>"
LOG_FORMAT_FILE = "{time:YYYY-MM-DD HH:mm:ss} | v{extra[ver]} | {name}:{line} | {level: <8} | {message}"

# 控制台输出
logger.add(
    lambda msg: print(msg, end=""),
    format=LOG_FORMAT_CONSOLE,
    level="INFO",
    colorize=True,
    filter=lambda record: not record["extra"].get("file_only")
)

# 文件日志（完整格式，包含代码位置）
log_file = LOGS_DIR / "veo3free.log"
logger.add(
    log_file,
    format=LOG_FORMAT_FILE,
    rotation="10 MB",
    retention="7 days",
    encoding="utf-8"
)

# 绑定版本号到所有日志
logger = logger.bind(ver=APP_VERSION)


def get_logger():
    """获取绑定了版本号的 logger"""
    return logger.bind(ver=APP_VERSION)


def log_error_to_file(message: str, exception: Exception = None):
    """记录错误：控制台显示ASCII安全版本，文件记录完整信息"""
    if exception:
        # 控制台用ascii转义，避免编码问题
        logger.error(f"{message}: {ascii(str(exception))}")
        # 文件记录完整堆栈
        logger.bind(file_only=True).exception(f"{message}")
    else:
        logger.error(message)


class TaskManager:
    """任务管理器"""

    TASK_TIMEOUT_SECONDS = 600  # 任务超时时间：10分钟
    CLIENT_COOLDOWN_SECONDS = 3  # 客户端冷却时间：3秒

    def __init__(self):
        self.tasks = []
        self.current_index = 0
        self.is_running = False
        self.clients = {}
        self.next_page_number = 1

    def register_client(self, websocket, page_url):
        import time
        for cid, info in list(self.clients.items()):
            if info['url'] == page_url:
                del self.clients[cid]

        client_id = f"c{len(self.clients)}_{int(time.time()) % 10000}"
        page_number = self.next_page_number
        self.next_page_number += 1
        self.clients[client_id] = {
            'ws': websocket,
            'url': page_url,
            'busy': False,
            'task_id': None,
            'page_number': page_number,
            'last_task_end': None
        }
        logger.info(f"客户端注册: {client_id} (页面{page_number})")
        return client_id, page_number

    def remove_client(self, client_id):
        if client_id in self.clients:
            page_number = self.clients[client_id].get('page_number')
            task_id = self.clients[client_id]['task_id']
            if task_id:
                for task in self.tasks:
                    if task['id'] == task_id and task['status'] == '处理中':
                        task['status'] = '等待中'
                        logger.warning(f"任务 {task_id} 因客户端断开重置为等待")
            del self.clients[client_id]
            logger.info(f"客户端断开: {client_id} (页面{page_number})")

    def get_idle_client(self):
        now = datetime.now()
        for cid, info in self.clients.items():
            if not info['busy']:
                # 检查冷却时间
                last_end = info.get('last_task_end')
                if last_end:
                    elapsed = (now - datetime.fromisoformat(last_end)).total_seconds()
                    if elapsed < self.CLIENT_COOLDOWN_SECONDS:
                        continue  # 还在冷却中，跳过这个客户端
                return cid, info
        return None, None

    def mark_client_busy(self, client_id, task_id):
        if client_id in self.clients:
            self.clients[client_id]['busy'] = True
            self.clients[client_id]['task_id'] = task_id
            for task in self.tasks:
                if task['id'] == task_id:
                    task['client_id'] = client_id
                    break

    def mark_client_idle(self, client_id):
        if client_id in self.clients:
            self.clients[client_id]['busy'] = False
            self.clients[client_id]['task_id'] = None
            self.clients[client_id]['last_task_end'] = datetime.now().isoformat()

    def get_client_count(self):
        total = len(self.clients)
        busy = sum(1 for c in self.clients.values() if c['busy'])
        return total, busy

    def update_task_status_detail(self, task_id, status_detail):
        for task in self.tasks:
            if task['id'] == task_id:
                task['status_detail'] = status_detail
                return True
        return False

    def add_task(self, prompt, task_type, aspect_ratio, resolution,
                 reference_images=None, output_dir=None, import_row_number=None):
        prompt = prompt.strip()
        if not prompt:
            return None

        if task_type == "Text to Video":
            reference_images = []

        task_id = f"task_{len(self.tasks)}_{datetime.now().strftime('%H%M%S%f')}"
        file_ext = ".mp4" if "Video" in task_type else ".png"

        task = {
            'id': task_id,
            'prompt': prompt,
            'status': '等待中',
            'status_detail': '',
            'file_ext': file_ext,
            'output_dir': output_dir,
            'client_id': None,
            'task_type': task_type,
            'aspect_ratio': aspect_ratio,
            'resolution': resolution,
            'reference_images': reference_images or [],
            'start_time': None,
            'end_time': None,
            'import_row_number': import_row_number  # 导入任务的行号（编号）
        }
        self.tasks.append(task)
        logger.info(f"添加任务: {task_id} | {task_type} | {aspect_ratio}")
        return task

    def get_next_task(self):
        while self.current_index < len(self.tasks):
            task = self.tasks[self.current_index]
            if task['status'] == '等待中':
                return task
            self.current_index += 1
        return None

    def check_timeout_tasks(self):
        """检查并处理超时任务，返回超时的任务列表"""
        timeout_tasks = []
        now = datetime.now()
        for task in self.tasks:
            if task['status'] == '处理中' and task.get('start_time'):
                start = datetime.fromisoformat(task['start_time'])
                elapsed = (now - start).total_seconds()
                if elapsed > self.TASK_TIMEOUT_SECONDS:
                    task['status'] = '超时'
                    task['end_time'] = now.isoformat()
                    task['status_detail'] = f'任务超时（超过{self.TASK_TIMEOUT_SECONDS // 60}分钟）'
                    logger.warning(f"任务超时: {task['id']} (耗时 {elapsed:.0f}s)")
                    # 释放对应客户端
                    client_id = task.get('client_id')
                    if client_id and client_id in self.clients:
                        self.clients[client_id]['busy'] = False
                        self.clients[client_id]['task_id'] = None
                    timeout_tasks.append(task)
        return timeout_tasks

    def clear_tasks(self):
        """清除所有任务"""
        self.tasks.clear()
        self.current_index = 0
        self.is_running = False
        logger.info("已清除所有任务")


class ImageProcessor:
    @staticmethod
    def compress_image_to_base64(image_path, max_size_bytes=768 * 1024):
        try:
            img = Image.open(image_path)
            if img.mode in ('RGBA', 'LA', 'P'):
                background = Image.new('RGB', img.size, (255, 255, 255))
                if img.mode == 'P':
                    img = img.convert('RGBA')
                background.paste(img, mask=img.split()[-1] if img.mode in ('RGBA', 'LA') else None)
                img = background
            elif img.mode != 'RGB':
                img = img.convert('RGB')

            quality = 95
            while quality > 5:
                buffer = io.BytesIO()
                img.save(buffer, format='JPEG', quality=quality, optimize=True)
                size = buffer.tell()
                if size <= max_size_bytes:
                    buffer.seek(0)
                    return base64.b64encode(buffer.getvalue()).decode('utf-8')
                quality -= 5

            scale = 0.9
            while scale > 0.1:
                new_size = (int(img.size[0] * scale), int(img.size[1] * scale))
                resized_img = img.resize(new_size, Image.Resampling.LANCZOS)
                buffer = io.BytesIO()
                resized_img.save(buffer, format='JPEG', quality=85, optimize=True)
                size = buffer.tell()
                if size <= max_size_bytes:
                    buffer.seek(0)
                    return base64.b64encode(buffer.getvalue()).decode('utf-8')
                scale -= 0.1

            buffer.seek(0)
            return base64.b64encode(buffer.getvalue()).decode('utf-8')
        except Exception as e:
            log_error_to_file("压缩图片失败", e)
            return None

    @staticmethod
    def generate_thumbnail(image_path, size=(200, 200)):
        """生成缩略图，返回 base64"""
        try:
            img = Image.open(image_path)
            # 转换为 RGB（处理 RGBA 等）
            if img.mode in ('RGBA', 'LA', 'P'):
                background = Image.new('RGB', img.size, (255, 255, 255))
                if img.mode == 'P':
                    img = img.convert('RGBA')
                background.paste(img, mask=img.split()[-1] if img.mode in ('RGBA', 'LA') else None)
                img = background
            elif img.mode != 'RGB':
                img = img.convert('RGB')

            # 生成缩略图
            img.thumbnail(size, Image.Resampling.LANCZOS)
            buffer = io.BytesIO()
            img.save(buffer, format='JPEG', quality=85, optimize=True)
            buffer.seek(0)
            return base64.b64encode(buffer.read()).decode('utf-8')
        except Exception as e:
            log_error_to_file("生成缩略图失败", e)
            return None


class ImageDownloader:
    @staticmethod
    async def save_base64_image(base64_data, filename, output_dir=None):
        if output_dir is None:
            output_dir = OUTPUT_DIR
        filepath = Path(output_dir) / filename
        try:
            image_data = base64.b64decode(base64_data)
            with open(filepath, 'wb') as f:
                f.write(image_data)
            # 返回绝对路径
            return filepath.absolute()
        except Exception as e:
            log_error_to_file("保存图片失败", e)
            return None


"""
说明：
- 原先的引导页服务（12346）与 WebSocket 服务（12343）已合并进 FastAPI（12346）
- 具体实现见：server/guide_routes.py 与 server/ws_routes.py
"""


class Api:
    """暴露给前端的 API"""

    def __init__(self, task_manager, loop=None):
        self.task_manager = task_manager
        self.loop = loop

    def add_task(self, prompt, task_type, aspect_ratio, resolution, reference_images, output_dir):
        task = self.task_manager.add_task(
            prompt=prompt,
            task_type=task_type,
            aspect_ratio=aspect_ratio,
            resolution=resolution,
            reference_images=reference_images or [],
            output_dir=output_dir or None
        )
        if task:
            logger.info(f"已添加任务，当前共 {len(self.task_manager.tasks)} 个")
            return {'success': True}
        return {'success': False, 'error': '添加失败'}

    def retry_task(self, task_index: int) -> dict:
        """重试失败的任务"""
        if task_index < 0 or task_index >= len(self.task_manager.tasks):
            return {'success': False, 'error': '任务索引无效'}
        
        task = self.task_manager.tasks[task_index]
        if task['status'] not in ('失败', '超时', '下载失败'):
            return {'success': False, 'error': '只能重试失败的任务'}
        
        # 重置任务状态
        task['status'] = '等待中'
        task['status_detail'] = ''
        task['start_time'] = None
        task['end_time'] = None
        task['client_id'] = None
        
        # 如果当前执行索引已经超过这个任务，需要回退
        if self.task_manager.current_index > task_index:
            self.task_manager.current_index = task_index
        
        logger.info(f"任务已重置为等待中: {task['id']}")
        
        # 如果执行已停止，自动启动
        if not self.task_manager.is_running:
            self.start_execution()
        
        return {'success': True}

    def retry_all_failed(self) -> dict:
        """重试所有失败的任务"""
        failed_indices = []
        for i, task in enumerate(self.task_manager.tasks):
            if task['status'] in ('失败', '超时', '下载失败'):
                failed_indices.append(i)
        
        if not failed_indices:
            return {'success': False, 'error': '没有失败的任务'}
        
        # 重置所有失败任务
        min_index = len(self.task_manager.tasks)
        for i in failed_indices:
            task = self.task_manager.tasks[i]
            task['status'] = '等待中'
            task['status_detail'] = ''
            task['start_time'] = None
            task['end_time'] = None
            task['client_id'] = None
            if i < min_index:
                min_index = i
        
        # 回退执行索引到最早的失败任务
        if self.task_manager.current_index > min_index:
            self.task_manager.current_index = min_index
        
        logger.info(f"已重置 {len(failed_indices)} 个失败任务")
        
        # 如果执行已停止，自动启动
        if not self.task_manager.is_running:
            self.start_execution()
        
        return {'success': True, 'count': len(failed_indices)}

    def clear_tasks(self) -> dict:
        """清除所有任务"""
        self.task_manager.clear_tasks()
        return {'success': True}

    def run_single_task(self, task_index: int) -> dict:
        """立即执行单个任务"""
        if task_index < 0 or task_index >= len(self.task_manager.tasks):
            return {'success': False, 'error': '任务索引无效'}
        
        task = self.task_manager.tasks[task_index]
        if task['status'] != '等待中':
            return {'success': False, 'error': '只能执行等待中的任务'}
        
        total, _ = self.task_manager.get_client_count()
        if total == 0:
            return {'success': False, 'error': '没有连接的客户端'}
        
        # 将任务移动到当前执行位置
        if task_index != self.task_manager.current_index:
            # 从原位置移除
            self.task_manager.tasks.pop(task_index)
            # 插入到当前执行位置
            self.task_manager.tasks.insert(self.task_manager.current_index, task)
        
        logger.info(f"立即执行任务: {task['id']}")
        
        # 启动执行
        if not self.task_manager.is_running:
            self.start_execution()
        
        return {'success': True}

    def get_status(self):
        total, busy = self.task_manager.get_client_count()
        tasks_data = []
        for t in self.task_manager.tasks:
            tasks_data.append({
                'id': t['id'],
                'prompt': t['prompt'],
                'status': t['status'],
                'status_detail': t.get('status_detail', ''),
                'task_type': t['task_type'],
                'aspect_ratio': t['aspect_ratio'],
                'resolution': t['resolution'],
                'saved_path': t.get('saved_path', ''),
                'output_dir': t.get('output_dir', ''),
                'start_time': t.get('start_time'),
                'end_time': t.get('end_time'),
                'file_ext': t.get('file_ext', ''),
                'preview_base64': t.get('preview_base64', '')
            })
        return {
            'client_count': total,
            'busy_count': busy,
            'is_running': self.task_manager.is_running,
            'tasks': tasks_data
        }

    def start_execution(self):
        if self.loop is None:
            logger.warning("启动执行失败: 后端事件循环未就绪")
            return
        total, _ = self.task_manager.get_client_count()
        if total == 0:
            logger.warning("启动执行失败: 没有连接的客户端")
            return
        if not self.task_manager.tasks:
            logger.warning("启动执行失败: 任务列表为空")
            return
        self.task_manager.is_running = True
        logger.info(f"启动任务执行: 客户端数={total}, 任务数={len(self.task_manager.tasks)}")
        try:
            running = asyncio.get_running_loop()
        except RuntimeError:
            running = None
        if running is not None and running is self.loop:
            asyncio.create_task(self._execute_tasks())
        else:
            asyncio.run_coroutine_threadsafe(self._execute_tasks(), self.loop)

    def stop_execution(self):
        self.task_manager.is_running = False
        logger.info("已停止执行")

    def switch_task_mode(self, task_type: str) -> dict:
        """向所有空闲客户端广播模式切换消息，让网页联动切换 Image/Video tab"""
        return self.sync_page_settings(task_type)

    def sync_page_settings(self, task_type: str, aspect_ratio: str = '16:9',
                           image_count: str = 'x1', image_model: str = 'Nano Banana 2') -> dict:
        """向所有空闲客户端广播完整设置同步消息（任务类型/比例/数量/模型）"""
        if self.loop is None:
            return {'success': False, 'error': '事件循环未就绪'}

        async def _broadcast():
            msg = json.dumps({
                'type': 'sync_settings',
                'task_type': task_type,
                'aspect_ratio': aspect_ratio,
                'image_count': image_count,
                'image_model': image_model,
            }, ensure_ascii=False)
            for _cid, info in list(self.task_manager.clients.items()):
                if not info.get('busy'):
                    try:
                        await info['ws'].send(msg)
                    except Exception:
                        pass

        asyncio.run_coroutine_threadsafe(_broadcast(), self.loop)
        logger.info(f"发送设置同步: type={task_type} ratio={aspect_ratio} count={image_count} model={image_model}")
        return {'success': True}

    async def _execute_tasks(self):
        logger.info("任务执行循环启动")

        while self.task_manager.is_running:
            # 检查超时任务
            timeout_tasks = self.task_manager.check_timeout_tasks()
            for t in timeout_tasks:
                logger.warning(f"任务超时: {t['id']}")

            task = self.task_manager.get_next_task()
            if not task:
                has_busy = any(c['busy'] for c in self.task_manager.clients.values())
                if not has_busy:
                    logger.info("所有任务已完成")
                    break
                await asyncio.sleep(1)
                continue

            # 检查是否有客户端连接
            total, busy = self.task_manager.get_client_count()
            if total == 0:
                # 没有客户端连接，等待客户端连接
                logger.info("没有客户端连接，等待客户端连接...")
                await asyncio.sleep(2)  # 等待时间稍长一些
                continue

            client_id, client_info = self.task_manager.get_idle_client()
            if not client_info:
                # 所有客户端都忙碌，等待空闲客户端
                await asyncio.sleep(1)
                continue

            task['status'] = '处理中'
            task['start_time'] = datetime.now().isoformat()
            self.task_manager.mark_client_busy(client_id, task['id'])
            self.task_manager.current_index += 1

            logger.info(f"分配任务: {task['id']} -> {client_id} | {task['task_type']}")

            # 延迟处理参考图片：如果是路径则压缩为 base64
            reference_images = []
            for img in task['reference_images']:
                if img and not img.startswith('/9j/') and not img.startswith('iVBOR') and Path(img).exists():
                    # 是文件路径，需要压缩
                    base64_data = ImageProcessor.compress_image_to_base64(img)
                    if base64_data:
                        reference_images.append(base64_data)
                else:
                    # 已经是 base64 数据
                    reference_images.append(img)

            message = json.dumps({
                'type': 'task',
                'task_id': task['id'],
                'prompt': task['prompt'],
                'task_type': task['task_type'],
                'aspect_ratio': task['aspect_ratio'],
                'resolution': task['resolution'],
                'reference_images': reference_images
            })

            try:
                await client_info['ws'].send(message)
            except Exception as e:
                logger.error(f"任务发送失败: {task['id']} -> {client_id}")
                log_error_to_file(f"任务发送失败", e)
                task['status'] = '等待中'
                self.task_manager.mark_client_idle(client_id)

            await asyncio.sleep(0.5)

        self.task_manager.is_running = False
        logger.info("任务队列执行结束")

    def select_images(self):
        """打开文件对话框选择图片"""
        file_types = ('图片文件 (*.png;*.jpg;*.jpeg;*.gif;*.bmp;*.webp)',)
        result = webview.windows[0].create_file_dialog(
            webview.OPEN_DIALOG,
            allow_multiple=True,
            file_types=file_types
        )
        if not result:
            return []

        images = []
        for filepath in result:
            logger.info(f"正在处理: {Path(filepath).name}")
            base64_data = ImageProcessor.compress_image_to_base64(filepath)
            if base64_data:
                images.append(base64_data)
                size_kb = len(base64_data) * 3 / 4 / 1024
                logger.info(f"已添加: {Path(filepath).name} (压缩后 ~{size_kb:.1f}KB)")
        return images

    def import_excel(self):
        """导入 Excel 文件"""
        if load_workbook is None:
            return {'success': False, 'count': 0, 'errors': ['请安装 openpyxl']}

        file_types = ('Excel文件 (*.xlsx)',)
        result = webview.windows[0].create_file_dialog(
            webview.OPEN_DIALOG,
            file_types=file_types
        )
        if not result:
            return {'success': False, 'count': 0, 'errors': []}

        filepath = result[0]
        logger.info(f"导入 Excel 开始: {Path(filepath).name}")

        # 验证分辨率和任务类型是否匹配
        def validate_resolution(task_type, resolution, aspect_ratio):
            """检查分辨率是否与任务类型兼容"""
            valid_resolutions = {
                "Create Image": ["4K", "2K", "1K"],
                "Text to Video": ["1080p", "720p"],
                "Frames to Video": ["1080p", "720p"],
                "Ingredients to Video": ["1080p", "720p"]
            }
            if task_type not in valid_resolutions:
                return False, f"未知任务类型: {task_type}"

            allowed = valid_resolutions[task_type]

            if resolution not in allowed:
                return False, f"{task_type} 不支持分辨率 {resolution}，请使用: {', '.join(allowed)}"
            return True, ""

        task_type_map = {
            "文生图片": "Create Image",
            "文生视频": "Text to Video",
            "图生视频": "Ingredients to Video",
            "首尾帧视频": "Frames to Video",
        }
        orientation_map = {
            "横屏": "16:9",
            "竖屏": "9:16"
        }

        try:
            wb = load_workbook(filepath)
            ws = wb.active

            # 第一步：验证所有行
            tasks_to_add = []
            validation_errors = []

            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not row or not row[1]:
                    continue

                try:
                    # 读取编号列（第一列）
                    row_number = str(row[0]).strip() if row[0] else str(row_idx)

                    prompt = str(row[1]).strip() if row[1] else ""
                    if not prompt:
                        continue

                    task_type_cn = str(row[2]).strip() if len(row) > 2 and row[2] else "图片"
                    orientation_cn = str(row[3]).strip() if len(row) > 3 and row[3] else "横屏"
                    resolution = str(row[4]).strip() if len(row) > 4 and row[4] else ""
                    output_dir = str(row[5]).strip() if len(row) > 5 and row[5] else None

                    # 验证任务类型
                    if task_type_cn not in task_type_map:
                        validation_errors.append(f"编号{row_number}: 未知任务类型: {task_type_cn}，请使用: {', '.join(task_type_map.keys())}")
                        continue

                    task_type = task_type_map[task_type_cn]
                    aspect_ratio = orientation_map.get(orientation_cn, "16:9")

                    if not resolution:
                        resolution = "1080p" if "Video" in task_type else "4K"
                    else:
                        # 分辨率忽略大小写处理
                        resolution_upper = resolution.upper()
                        resolution_lower = resolution.lower()
                        # 标准化为正确的格式（如 4k -> 4K, 1080p -> 1080p）
                        if resolution_upper in ["4K", "2K", "1K"]:
                            resolution = resolution_upper
                        elif resolution_lower == "1080p":
                            resolution = "1080p"
                        elif resolution_lower == "720p":
                            resolution = "720p"

                    # 验证分辨率
                    is_valid, error_msg = validate_resolution(task_type, resolution, aspect_ratio)
                    if not is_valid:
                        validation_errors.append(f"行{row_idx}: {error_msg}")
                        continue

                    reference_images = []
                    max_images = {
                        "Create Image": 8,
                        "Frames to Video": 2,
                        "Ingredients to Video": 3,
                        "Text to Video": 0
                    }.get(task_type, 8)

                    # 只收集图片路径，不在导入时压缩（延迟到执行时处理）
                    for i in range(max_images):
                        col_idx = 6 + i
                        if len(row) > col_idx and row[col_idx]:
                            img_path = str(row[col_idx]).strip()
                            if img_path and Path(img_path).exists():
                                reference_images.append(img_path)

                    tasks_to_add.append({
                        'prompt': prompt,
                        'task_type': task_type,
                        'aspect_ratio': aspect_ratio,
                        'resolution': resolution,
                        'reference_images': reference_images,
                        'output_dir': output_dir,
                        'import_row_number': row_number  # Excel 编号列的值
                    })

                except Exception as e:
                    validation_errors.append(f"编号{row_number}: {str(e)}")

            wb.close()

            # 如果有验证错误，全部不导入
            if validation_errors:
                return {'success': False, 'count': 0, 'errors': validation_errors}

            # 如果没有任何有效任务
            if not tasks_to_add:
                return {'success': False, 'count': 0, 'errors': ['没有找到有效的任务行']}

            # 第二步：全部验证通过，一次性导入所有任务
            for task_data in tasks_to_add:
                self.task_manager.add_task(
                    prompt=task_data['prompt'],
                    task_type=task_data['task_type'],
                    aspect_ratio=task_data['aspect_ratio'],
                    resolution=task_data['resolution'],
                    reference_images=task_data['reference_images'],
                    output_dir=task_data['output_dir'],
                    import_row_number=task_data['import_row_number']
                )

            count = len(tasks_to_add)
            logger.info(f"从Excel导入 {count} 个任务")
            return {'success': True, 'count': count, 'errors': []}

        except Exception as e:
            return {'success': False, 'count': 0, 'errors': [str(e)]}

    def export_template(self):
        """导出 Excel 模板"""
        if Workbook is None:
            return

        file_types = ('Excel文件 (*.xlsx)',)
        result = webview.windows[0].create_file_dialog(
            webview.SAVE_DIALOG,
            file_types=file_types,
            save_filename='高级模板.xlsx'
        )
        if not result:
            return

        filepath = result if isinstance(result, str) else result[0]

        # 确保有 .xlsx 扩展名
        if not filepath.lower().endswith('.xlsx'):
            filepath += '.xlsx'

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "任务列表"

            headers = ["编号", "提示词", "任务类型", "屏幕方向", "分辨率", "输出文件夹",
                       "图1", "图2", "图3", "图4", "图5", "图6", "图7", "图8"]
            for col, header in enumerate(headers, start=1):
                ws.cell(row=1, column=col, value=header)

            examples = [
                [1, "A beautiful sunset over the ocean", "文生图片", "横屏", "4K", "sunset"],
                [2, "A beautiful moon over the ocean", "文生图片", "竖屏", "2K", "sunset"],
                [3, "A cute cat playing", "文生视频", "横屏", "1080p", "cats"],
                [4, "A cute dog playing", "文生视频", "竖屏", "720p", "dogs_注意veo3竖屏视频不支持1080p"],
                [5, "动起来", "首尾帧视频", "横屏", "1080p", "frames", "/Users/wei/Downloads/pig.jpeg"],
                [6, "组合这些照片为一个创意视频", "图生视频", "横屏", "1080p", "collage", "/Users/wei/Downloads/pig.jpeg"],
            ]

            for row_idx, example in enumerate(examples, start=2):
                for col_idx, value in enumerate(example, start=1):
                    ws.cell(row=row_idx, column=col_idx, value=value)

            wb.save(filepath)
            logger.info(f"已导出高级模板: {filepath}")

        except Exception as e:
            log_error_to_file("导出模板失败", e)

    def open_output_dir(self):
        """打开输出目录"""
        self._open_directory(OUTPUT_DIR)

    def open_logs_dir(self):
        """打开日志目录"""
        self._open_directory(LOGS_DIR)

    def open_task_file(self, task_index):
        """打开任务的文件（不是文件夹）"""
        if 0 <= task_index < len(self.task_manager.tasks):
            task = self.task_manager.tasks[task_index]
            saved_path = task.get('saved_path')
            if saved_path and Path(saved_path).exists():
                path = Path(saved_path)
                # 直接打开文件
                path_str = str(path.absolute())
                system = platform.system()

                if system == 'Windows':
                    os.startfile(path_str)
                elif system == 'Darwin':
                    subprocess.run(['open', path_str])
                else:
                    subprocess.run(['xdg-open', path_str])
            else:
                # 文件不存在，打开所在目录
                self.open_task_dir(task_index)
        else:
            self.open_output_dir()

    def open_task_dir(self, task_index):
        """打开任务的输出目录"""
        if 0 <= task_index < len(self.task_manager.tasks):
            task = self.task_manager.tasks[task_index]
            output_dir = task.get('output_dir_path', str(OUTPUT_DIR))
            self._open_directory(Path(output_dir))
        else:
            self._open_directory(OUTPUT_DIR)

    def _open_directory(self, path):
        path = Path(path)
        if not path.exists():
            path.mkdir(parents=True, exist_ok=True)

        path_str = str(path.absolute())
        system = platform.system()

        if system == 'Windows':
            os.startfile(path_str)
        elif system == 'Darwin':
            subprocess.run(['open', path_str])
        else:
            subprocess.run(['xdg-open', path_str])

    def get_app_version(self) -> str:
        """获取应用版本号"""
        version = get_version()
        logger.info(f"获取应用版本: {version}")
        return version

    def check_update(self) -> dict:
        """检查更新"""
        logger.info("前端请求检查更新")
        info = check_for_updates()

        if info is None:
            logger.warning("更新检查返回 None，检查失败")
            return {
                'success': False,
                'has_update': False,
                'current_version': get_version(),
                'latest_version': '',
                'release_notes': '',
                'download_url': '',
                'release_url': ''
            }

        logger.info(f"更新检查完成: 有更新={info.has_update}, 最新版本={info.latest_version}")
        return {
            'success': True,
            'has_update': info.has_update,
            'current_version': info.current_version,
            'latest_version': info.latest_version,
            'release_notes': info.release_notes,
            'download_url': info.download_url,
            'release_url': info.release_url
        }

    def open_update_page(self, url: str) -> bool:
        """在浏览器中打开更新下载页面"""
        logger.info(f"打开下载页面: {url}")
        return open_download_page(url)

    def open_guide_page(self) -> bool:
        """在外部浏览器中打开引导页面"""
        guide_url = f"http://localhost:{API_SERVER_PORT}/guide"
        logger.info(f"打开引导页面: {guide_url}")
        try:
            webbrowser.open(guide_url)
            return True
        except Exception as e:
            logger.error(f"打开引导页面失败: {e}")
            return False

    def _get_settings_path(self) -> Path:
        """获取配置文件路径：~/veo3free/settings.json"""
        return Path.home() / "veo3free" / "settings.json"

    def _load_settings(self) -> dict:
        """加载配置文件"""
        settings_path = self._get_settings_path()
        if settings_path.exists():
            try:
                return json.loads(settings_path.read_text(encoding="utf-8"))
            except Exception:
                pass
        return {}

    def _save_settings(self, settings: dict) -> bool:
        """保存配置文件"""
        settings_path = self._get_settings_path()
        try:
            settings_path.parent.mkdir(parents=True, exist_ok=True)
            settings_path.write_text(json.dumps(settings, ensure_ascii=False, indent=2), encoding="utf-8")
            return True
        except Exception as e:
            logger.warning(f"保存配置文件失败: {e}")
            return False

    def _get_persistent_device_id(self) -> str:
        """获取持久化的设备 ID，首次启动时生成随机 UUID 并保存"""
        import uuid

        settings = self._load_settings()

        # 已有 device_id 则直接返回
        if settings.get("device_id"):
            return settings["device_id"]

        # 首次启动，生成随机 UUID
        device_id = str(uuid.uuid4()).replace("-", "")
        settings["device_id"] = device_id
        self._save_settings(settings)

        return device_id

    def _get_or_create_api_key(self) -> str:
        """获取或创建 API 密钥"""
        import secrets
        settings = self._load_settings()
        
        if settings.get("api_key"):
            return settings["api_key"]
        
        # 生成新的 API 密钥
        api_key = "vf-" + secrets.token_urlsafe(24)
        settings["api_key"] = api_key
        self._save_settings(settings)
        return api_key

    def get_api_verify_status(self) -> dict:
        """获取 API 验证状态"""
        settings = self._load_settings()
        if settings.get("api_verified"):
            return {
                'verified': True,
                'api_key': settings.get("api_key"),
                'docs_url': f'http://localhost:{API_SERVER_PORT}/docs'
            }
        return {'verified': False}

    def verify_pjy_card(self, card: str) -> dict:
        """验证泡椒云卡密"""
        if not PJYSDK_AVAILABLE:
            return {
                'success': False,
                'error': 'PJYSDK 未安装，验证功能不可用'
            }

        try:
            import uuid
            import hashlib

            app_key = 'd5rg9r3dqussfksagchg'
            app_secret = 'hCDOnEu6G66CThuxJG2BMXmjfN9wGRQ6'

            pjysdk = PJYSDK(app_key=app_key, app_secret=app_secret)
            pjysdk.debug = False

            device_id = self._get_persistent_device_id()
            pjysdk.set_device_id(device_id)
            pjysdk.set_card(card)

            ret = pjysdk.card_login()
            if ret.code == 0:
                token = ret.result.token
                logger.info(f"验证成功: token={token[:20]}...")
                
                # 保存验证状态
                settings = self._load_settings()
                settings["api_card"] = card
                settings["api_verified"] = True
                settings["api_token"] = token
                # 确保有 api_key
                if not settings.get("api_key"):
                    import secrets
                    settings["api_key"] = "vf-" + secrets.token_urlsafe(24)
                self._save_settings(settings)
                
                return {
                    'success': True,
                    'token': token,
                    'api_key': settings["api_key"],
                    'docs_url': f'http://localhost:{API_SERVER_PORT}/docs'
                }
            else:
                logger.warning(f"泡椒云验证失败: {ret.message}")
                return {
                    'success': False,
                    'error': ret.message or '验证失败'
                }
        except Exception as e:
            logger.error(f"泡椒云验证异常: {e}")
            log_error_to_file("泡椒云验证异常", e)
            return {
                'success': False,
                'error': f'验证异常: {str(e)}'
            }

    def select_image_folder(self) -> dict:
        """选择文件夹并扫描图片，返回图片路径列表"""
        try:
            # 从用户主目录开始选择
            start_dir = str(Path.home())
            result = webview.windows[0].create_file_dialog(
                webview.FOLDER_DIALOG,
                directory=start_dir
            )
            if not result:
                return {'success': False, 'folder_path': '', 'images': []}

            folder_path = result[0] if isinstance(result, (list, tuple)) else result
            folder = Path(folder_path)

            if not folder.exists() or not folder.is_dir():
                return {'success': False, 'folder_path': '', 'images': [], 'error': '选择的路径不是有效文件夹'}

            # 扫描图片文件
            image_extensions = {'.jpg', '.jpeg', '.png', '.webp'}
            images = []
            for file in sorted(folder.iterdir()):
                if file.is_file() and file.suffix.lower() in image_extensions:
                    images.append(str(file.absolute()))

            logger.info(f"扫描文件夹: {folder_path}, 找到 {len(images)} 张图片")
            return {
                'success': True,
                'folder_path': str(folder_path),
                'images': images
            }
        except Exception as e:
            log_error_to_file("选择图片文件夹失败", e)
            return {'success': False, 'folder_path': '', 'images': [], 'error': str(e)}

    def create_custom_template(self, images: list, task_type: str, aspect_ratio: str,
                                resolution: str, output_dir: str, default_prompt: str) -> dict:
        """根据图片列表和参数创建预填充的 Excel 模板"""
        if Workbook is None:
            return {'success': False, 'error': '请安装 openpyxl'}

        if not images:
            return {'success': False, 'error': '没有图片可导出'}

        # 任务类型映射（英文 -> 中文）
        task_type_map = {
            "Create Image": "文生图片",
            "Frames to Video": "首尾帧视频",
            "Ingredients to Video": "图生视频",
        }

        # 屏幕方向映射
        orientation_map = {
            "16:9": "横屏",
            "9:16": "竖屏"
        }

        task_type_cn = task_type_map.get(task_type, task_type)
        orientation_cn = orientation_map.get(aspect_ratio, aspect_ratio)

        file_types = ('Excel文件 (*.xlsx)',)
        # 从用户文档目录开始保存
        start_dir = str(Path.home() / "Documents")
        result = webview.windows[0].create_file_dialog(
            webview.SAVE_DIALOG,
            directory=start_dir,
            file_types=file_types,
            save_filename='简单模板.xlsx'
        )
        if not result:
            return {'success': False, 'error': '取消保存'}

        filepath = result if isinstance(result, str) else result[0]

        # 确保有 .xlsx 扩展名
        if not filepath.lower().endswith('.xlsx'):
            filepath += '.xlsx'

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "任务列表"

            # 表头
            headers = ["编号", "提示词", "任务类型", "屏幕方向", "分辨率", "输出文件夹",
                       "图1", "图2", "图3", "图4", "图5", "图6", "图7", "图8"]
            for col, header in enumerate(headers, start=1):
                ws.cell(row=1, column=col, value=header)

            # 为每张图片生成一行
            for idx, img_path in enumerate(images, start=1):
                row_idx = idx + 1  # 第2行开始
                ws.cell(row=row_idx, column=1, value=idx)  # 编号
                ws.cell(row=row_idx, column=2, value=default_prompt or "")  # 默认提示词
                ws.cell(row=row_idx, column=3, value=task_type_cn)  # 任务类型
                ws.cell(row=row_idx, column=4, value=orientation_cn)  # 屏幕方向
                ws.cell(row=row_idx, column=5, value=resolution)  # 分辨率
                ws.cell(row=row_idx, column=6, value=output_dir or "")  # 输出文件夹
                ws.cell(row=row_idx, column=7, value=img_path)  # 图1

            wb.save(filepath)
            logger.info(f"已创建简单模板: {filepath}, 包含 {len(images)} 行")
            return {'success': True, 'filepath': filepath, 'count': len(images)}

        except Exception as e:
            log_error_to_file("创建专用模板失败", e)
            return {'success': False, 'error': str(e)}
def main():
    # 启动日志
    logger.info("=" * 50)
    logger.info(f"Veo3Free 启动 - 版本: {get_version()}")
    logger.info(f"运行环境: {'打包模式' if getattr(sys, 'frozen', False) else '开发模式'}")
    logger.info(f"操作系统: {platform.system()} {platform.release()}")
    logger.info(f"Python: {sys.version}")
    logger.info(f"输出目录: {OUTPUT_DIR}")
    logger.info(f"日志目录: {LOGS_DIR}")
    logger.info("=" * 50)

    # 创建任务管理器和 API
    task_manager = TaskManager()
    api = Api(task_manager)

    # 获取或创建 API 密钥
    api_key = api._get_or_create_api_key()

    # 启动 OpenAI API 兼容服务器（同时承载 guide 与 WebSocket）
    logger.info(f"正在启动 OpenAI API 服务器 (端口 {API_SERVER_PORT})...")
    try:
        logger.info(f"API server module: {Path(api_server_module.__file__).resolve()}")
    except Exception:
        logger.info("API server module: <unknown>")
    api_server = OpenAIAPICompatServer(
        task_manager, output_dir=OUTPUT_DIR, port=API_SERVER_PORT, api_key=api_key, api_instance=api
    )
    # 让 Api 可以访问 api_server
    api._api_server = api_server
    if api_server.start():
        logger.info(f"OpenAI API 服务器启动成功: http://localhost:{API_SERVER_PORT}")
        logger.info(f"文件访问端点: http://localhost:{API_SERVER_PORT}/files/<task_id>")
        logger.info(f"引导页面: http://localhost:{API_SERVER_PORT}/guide")
        logger.info(f"WebSocket: ws://localhost:{API_SERVER_PORT}/ws")
    else:
        logger.warning("OpenAI API 服务器启动失败")

    # 等待 FastAPI loop 就绪后，注入到 Api 里用于任务调度
    if api_server.wait_ready(timeout_seconds=5.0) and api_server.loop is not None:
        api.loop = api_server.loop
    else:
        logger.warning("FastAPI 事件循环未就绪，任务调度可能不可用")

    # 确定 web 目录和 URL
    if getattr(sys, 'frozen', False):
        # 打包后，使用打包的web目录
        web_dir = Path(sys._MEIPASS) / 'web'
        url = str(web_dir / 'index.html')
        logger.info(f"使用打包资源: {url}")
    else:
        # 开发模式
        web_dir = Path(__file__).parent / 'web'
        # 检查是否使用开发服务器
        if os.environ.get('DEV') == '1' or not web_dir.exists():
            url = 'http://localhost:9173'
            logger.info("使用开发服务器: http://localhost:9173")
        else:
            url = str(web_dir / 'index.html')
            logger.info(f"使用本地文件: {url}")

    # 创建窗口
    logger.info("正在创建应用窗口...")
    window = webview.create_window(
        'Veo3Free - AI生成工具',
        url,
        width=1000,
        height=700,
        min_size=(800, 600),
        maximized=True,
        js_api=api
    )
    # !!! 严谨对api设置window等对象，例如"api.window = window"是极其危险的！！！

    # 启动 webview
    logger.info("启动 webview 主循环...")
    webview.start(gui='qt')

    # 清理
    logger.info("正在关闭应用...")
    api_server.stop()
    logger.info("应用已退出")


if __name__ == "__main__":
    main()
